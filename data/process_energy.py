import json
import re
import sys
from collections import defaultdict
from datetime import date

import openpyxl

from la_config import ALL_LAS, HAMPSHIRE_SOLENT_LAS, MID_HAMPSHIRE_LAS, MID_HAMPSHIRE_RETAINED_FRACTION

RENEWABLE_SRC = sys.argv[1] if len(sys.argv) > 1 else "renewable_electricity_source.xlsx"
TFEC_SRC = sys.argv[2] if len(sys.argv) > 2 else "energy_consumption_source.xlsx"

# DUKES/IEA standard conversion, also used by DESNZ itself throughout this dataset family.
KTOE_TO_MWH = 11630.0

# The TFEC workbook's column headers wrap onto several lines and carry footnote markers that
# shift year to year (e.g. "Coal:\nTotal\n[Note 2]") — normalize to a stable "Coal: Total" form
# before matching, rather than relying on the exact (fragile) header text.
def normalize_header(h):
    if not h:
        return h
    h = re.sub(r"\[.*?\]", "", str(h))
    return re.sub(r"\s+", " ", h).strip()


# DESNZ's own fuel categories (in the "energy sources" chart's "complex" view). No suppression
# in this dataset (unlike renewable generation above), so these are used as published.
FUEL_CATEGORIES = ["Coal", "Manufactured fuels", "Petroleum", "Gas", "Electricity", "Bioenergy and wastes"]

# Groups the workbook's 12 raw technology columns into a handful of chart-friendly buckets.
# Wave/Tidal is folded into "Other" rather than "Wind" since it's a distinct (marine, not wind)
# technology that's negligible in these local authorities. "Other" also absorbs the gap between
# the sum of the visible columns and the workbook's own "Total" column, which is where DESNZ's
# small-plant disclosure suppression (cells marked "[X]") ends up — see note_suppression below.
TECH_GROUPS = {
    "Solar": ["Photovoltaics"],
    "Wind": ["Onshore Wind", "Offshore Wind"],
    "Hydro": ["Hydro"],
    "Bioenergy & waste": [
        "Anaerobic Digestion", "Sewage Gas", "Landfill Gas", "Municipal Solid Waste",
        "Animal Biomass", "Plant Biomass", "Cofiring",
    ],
    "Other": ["Wave/Tidal"],
}
TECH_GROUP_NAMES = list(TECH_GROUPS.keys())


def numeric(v):
    return v if isinstance(v, (int, float)) else 0.0


def read_generation_year(wb, year):
    sheet_name = f"LA - Generation, {year}"
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    header = None
    by_la = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "Local Authority Code [note 1]":
                header = row
            continue
        la = row[1]
        if la not in ALL_LAS:
            continue
        cols = {header[i]: row[i] for i in range(4, len(header)) if header[i]}
        total = numeric(cols.get("Total"))
        groups = {}
        accounted = 0.0
        for group, raw_cols in TECH_GROUPS.items():
            v = sum(numeric(cols.get(c)) for c in raw_cols)
            groups[group] = v
            accounted += v
        # Suppressed ("[X]") cells are treated as 0 above; fold the (small) residual against the
        # workbook's own Total into "Other" so group totals still sum exactly to DESNZ's figure.
        groups["Other"] += max(0.0, total - accounted)
        by_la[la] = {"total_mwh": total, "by_technology_mwh": groups}
    return by_la


def read_tfec_year(wb, year):
    sheet_name = str(year)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    header = None
    by_la = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "Code":
                header = [normalize_header(h) for h in row]
            continue
        la = row[2]
        if la not in ALL_LAS:
            continue
        cols = {header[i]: row[i] for i in range(len(header)) if header[i]}
        fuels_ktoe = {f: numeric(cols.get(f"{f}: Total")) for f in FUEL_CATEGORIES}
        by_la[la] = {
            "electricity_consumption_mwh": fuels_ktoe["Electricity"] * KTOE_TO_MWH,
            "fuels_ktoe": fuels_ktoe,
            "all_fuels_ktoe": numeric(cols.get("All fuels: Total")),
        }
    return by_la


def build_region_series(per_year_la, la_list, la_weight, value_key):
    weight = lambda la: (la_weight[la] if la_weight else 1.0)
    out = {}
    for year, by_la in per_year_la.items():
        if by_la is None:
            continue
        present = [la for la in la_list if la in by_la]
        if not present:
            continue
        if value_key == "generation":
            total = sum(by_la[la]["total_mwh"] * weight(la) for la in present)
            groups = {
                g: round(sum(by_la[la]["by_technology_mwh"][g] * weight(la) for la in present), 3)
                for g in TECH_GROUP_NAMES
            }
            out[year] = {"total_mwh": round(total, 3), "by_technology_mwh": groups}
        else:
            total = sum(by_la[la]["electricity_consumption_mwh"] * weight(la) for la in present)
            fuels = {
                f: round(sum(by_la[la]["fuels_ktoe"][f] * weight(la) for la in present), 4)
                for f in FUEL_CATEGORIES
            }
            all_fuels = round(sum(by_la[la]["all_fuels_ktoe"] * weight(la) for la in present), 4)
            out[year] = {
                "electricity_consumption_mwh": round(total, 3),
                "fuels_ktoe": fuels,
                "all_fuels_ktoe": all_fuels,
            }
    return out


def main():
    renewable_wb = openpyxl.load_workbook(RENEWABLE_SRC, read_only=True, data_only=True)
    tfec_wb = openpyxl.load_workbook(TFEC_SRC, read_only=True, data_only=True)

    generation_by_year = {}
    for year in range(2014, 2031):
        by_la = read_generation_year(renewable_wb, year)
        if by_la is not None:
            generation_by_year[year] = by_la

    consumption_by_year = {}
    for year in range(2005, 2031):
        by_la = read_tfec_year(tfec_wb, year)
        if by_la is not None:
            consumption_by_year[year] = by_la

    generation_years = sorted(generation_by_year.keys())
    consumption_years = sorted(consumption_by_year.keys())

    regions_spec = [
        ("winchester", "Winchester", ["Winchester"], None),
        ("mid-hampshire", "Mid-Hampshire (proposed)", MID_HAMPSHIRE_LAS, MID_HAMPSHIRE_RETAINED_FRACTION),
        ("hampshire-solent", "Hampshire and the Solent", HAMPSHIRE_SOLENT_LAS, None),
    ]

    out = {
        "meta": {
            "source_renewable": "DESNZ Regional renewable statistics — renewable electricity by local authority",
            "source_renewable_url": "https://www.gov.uk/government/statistics/regional-renewable-statistics",
            "source_consumption": "DESNZ Total final energy consumption at regional and local authority level",
            "source_consumption_url": "https://www.gov.uk/government/collections/total-final-energy-consumption-at-sub-national-level",
            "units": "MWh",
            "generation_years": generation_years,
            "consumption_years": consumption_years,
            "technology_groups": TECH_GROUP_NAMES,
            "fuel_categories": FUEL_CATEGORIES,
            "units_consumption": "ktoe (kilotonnes of oil equivalent), except electricity_consumption_mwh which is MWh",
            "note_boundary": "Same Mid-Hampshire / Hampshire and the Solent constituent local authorities and Mid-Hampshire population-based retained fractions as mid_hampshire_emissions.json — see that file's note_boundary for the full explanation.",
            "note_suppression": "DESNZ suppresses some small per-technology generation cells (marked \"[X]\" in the source workbook) to avoid revealing individual plants' output. This site treats suppressed cells as 0 for their own technology group and adds the (small) gap between the visible columns and DESNZ's own published Total into the \"Other\" group, so technology totals always sum exactly to DESNZ's published local authority total. The consumption-by-fuel dataset has no equivalent suppression.",
            "note_ktoe_conversion": "Energy consumption is published in ktoe (kilotonnes of oil equivalent); electricity_consumption_mwh converts the Electricity fuel category to MWh using the standard DUKES/IEA factor of 1 toe = 11.63 MWh, for comparison against renewable generation (also in MWh). The consumption-by-fuel chart displays all fuels in ktoe, DESNZ's native unit.",
            "generated": date.today().isoformat(),
        },
        "regions": {},
    }

    for key, name, la_list, weight in regions_spec:
        out["regions"][key] = {
            "name": name,
            "generation": build_region_series(generation_by_year, la_list, weight, "generation"),
            "consumption": build_region_series(consumption_by_year, la_list, weight, "consumption"),
        }

    with open("mid_hampshire_energy.json", "w") as f:
        json.dump(out, f, indent=1)

    with open("mid_hampshire_energy.js", "w") as f:
        f.write("window.MHE_ENERGY_DATA = ")
        json.dump(out, f, indent=1)
        f.write(";\n")

    print("Generation years:", generation_years[0], "-", generation_years[-1])
    print("Consumption years:", consumption_years[0], "-", consumption_years[-1])
    latest_shared_year = min(generation_years[-1], consumption_years[-1])
    for key, _name, _la_list, _weight in regions_spec:
        r = out["regions"][key]
        gen = r["generation"][latest_shared_year]["total_mwh"]
        con = r["consumption"][latest_shared_year]["electricity_consumption_mwh"]
        print(f"{key} {latest_shared_year}: renewable generation {gen:,.0f} MWh = "
              f"{100 * gen / con:.1f}% of electricity consumption ({con:,.0f} MWh)")


if __name__ == "__main__":
    main()
