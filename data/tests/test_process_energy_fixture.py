"""
End-to-end fixture test for process_energy.py — runs the real script against tiny synthetic
xlsx workbooks (same shape as DESNZ's real renewable-generation, TFEC and DUKES 6.5 releases),
including a disclosure-suppressed ("[X]") cell, to check the technology-grouping /
suppression-folding / region-rollup / DUKES-parsing logic without needing the real source
workbooks (the renewable and TFEC ones are ~4MB each; DUKES 6.5 is much smaller but still a
real download).
"""
import json
import os
import subprocess
import sys
import unittest
from tempfile import TemporaryDirectory

import openpyxl

DATA_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, DATA_DIR)

import la_config as lc  # noqa: E402

YEAR = 2020

GENERATION_HEADER = [
    "Local Authority Code [note 1]", "Local Authority", "Region Code", "Region",
    "Photovoltaics", "Onshore Wind", "Offshore Wind", "Hydro",
    "Anaerobic Digestion", "Sewage Gas", "Landfill Gas", "Municipal Solid Waste",
    "Animal Biomass", "Plant Biomass", "Cofiring", "Wave/Tidal", "Total",
]

TFEC_HEADER = [
    "Code", "Region", "Local Authority",
    "Coal: Total", "Manufactured fuels: Total", "Petroleum: Total", "Gas: Total",
    "Electricity: Total", "Bioenergy and wastes: Total", "All fuels: Total",
    "All fuels: Domestic", "All fuels: Transport", "All fuels: Industrial, Commercial and other",
]


def build_generation_workbook(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(f"LA - Generation, {YEAR}")
    ws.append(GENERATION_HEADER)
    # Winchester: fully disclosed, no suppression. Solar=100, OnshoreWind=50, everything else 0,
    # Total = exactly the visible sum (200 = 100+50+... wait keep simple: Total=150).
    ws.append(["E1", "Winchester", "R1", "South East", 100, 50, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 150])
    # East Hampshire: one column suppressed ("[X]") and Total exceeds the visible sum by the
    # suppressed plant's real (undisclosed) output — must land entirely in "Other".
    ws.append(["E2", "East Hampshire", "R1", "South East", 40, "[X]", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 65])
    for la in lc.ALL_LAS:
        if la in ("Winchester", "East Hampshire"):
            continue
        ws.append([f"E{la}", la, "R1", "South East", 10, 5, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 15])
    wb.save(path)


def build_dukes_workbook(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("6.5a")
    # Same shape as the real DUKES 6.5 workbook's "6.5a" sheet: a header row of years (as text,
    # matching the real workbook), then several unrelated rows before the one this app actually
    # reads, to check the parser locates it by label rather than by a fixed row offset.
    ws.append(["Electricity generation (ktoe)", str(YEAR - 1), str(YEAR)])
    ws.append(["Renewable generation", 400.0, 500.0])
    ws.append(["Total generation", 2000.0, 2500.0])
    ws.append(["Share of renewable generation", 0.19, 0.5075530080733758])
    wb.save(path)


def build_tfec_workbook(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(str(YEAR))
    ws.append(TFEC_HEADER)
    for la in lc.ALL_LAS:
        # Coal, Manufactured fuels, Petroleum, Gas, Electricity, Bioenergy and wastes, All fuels
        # (1+2+3+4+5+6=21), then the same 21 total split a different way by sector
        # (Domestic=7, Transport=8, Industrial/Commercial/other=6) — same total, different axis,
        # matching the two real columns groups process_energy.py reads from the TFEC workbook.
        ws.append(["C" + la, "South East", la, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 21.0, 7.0, 8.0, 6.0])
    wb.save(path)


class TestProcessEnergyFixture(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = TemporaryDirectory()
        gen_path = os.path.join(cls.tmpdir.name, "gen.xlsx")
        tfec_path = os.path.join(cls.tmpdir.name, "tfec.xlsx")
        dukes_path = os.path.join(cls.tmpdir.name, "dukes.xlsx")
        build_generation_workbook(gen_path)
        build_tfec_workbook(tfec_path)
        build_dukes_workbook(dukes_path)

        env = dict(os.environ, PYTHONPATH=DATA_DIR)
        result = subprocess.run(
            [sys.executable, os.path.join(DATA_DIR, "process_energy.py"), gen_path, tfec_path, dukes_path],
            cwd=cls.tmpdir.name, env=env, capture_output=True, text=True,
        )
        cls.process_result = result
        with open(os.path.join(cls.tmpdir.name, "mid_hampshire_energy.json")) as f:
            cls.out = json.load(f)

    @classmethod
    def tearDownClass(cls):
        cls.tmpdir.cleanup()

    def test_process_energy_py_exits_cleanly(self):
        self.assertEqual(self.process_result.returncode, 0, self.process_result.stderr)

    def test_suppressed_cell_folds_into_other_and_total_still_matches(self):
        gen = self.out["regions"]["east-hampshire"]["generation"][str(YEAR)]
        self.assertEqual(gen["total_mwh"], 65.0)
        # Visible Solar=40, suppressed Onshore Wind treated as 0 -> gap of 65-40=25 folds into Other.
        self.assertAlmostEqual(gen["by_technology_mwh"]["Solar"], 40.0, places=3)
        self.assertAlmostEqual(gen["by_technology_mwh"]["Other"], 25.0, places=3)
        total = sum(gen["by_technology_mwh"].values())
        self.assertAlmostEqual(total, gen["total_mwh"], places=3)

    def test_unsuppressed_technology_totals_exact(self):
        gen = self.out["regions"]["winchester"]["generation"][str(YEAR)]
        self.assertAlmostEqual(gen["by_technology_mwh"]["Solar"], 100.0, places=3)
        self.assertAlmostEqual(gen["by_technology_mwh"]["Wind"], 50.0, places=3)
        self.assertAlmostEqual(gen["by_technology_mwh"]["Other"], 0.0, places=3)

    def test_weighted_region_generation_scales_by_retained_fraction(self):
        year = str(YEAR)
        expected = sum(
            15.0 * lc.MID_HAMPSHIRE_RETAINED_FRACTION[la] if la != "Winchester" else 150.0 * lc.MID_HAMPSHIRE_RETAINED_FRACTION[la]
            for la in lc.MID_HAMPSHIRE_LAS
        )
        # East Hampshire is in MID_HAMPSHIRE_LAS too, and it uses 65 not 15/150 — recompute properly.
        totals = {"Winchester": 150.0, "East Hampshire": 65.0, "New Forest": 15.0, "Test Valley": 15.0}
        expected = sum(totals[la] * lc.MID_HAMPSHIRE_RETAINED_FRACTION[la] for la in lc.MID_HAMPSHIRE_LAS)
        actual = self.out["regions"]["mid-hampshire"]["generation"][year]["total_mwh"]
        self.assertAlmostEqual(actual, expected, places=2)

    def test_electricity_consumption_ktoe_conversion(self):
        c = self.out["regions"]["winchester"]["consumption"][str(YEAR)]
        self.assertAlmostEqual(c["fuels_ktoe"]["Electricity"], 5.0, places=3)
        self.assertAlmostEqual(c["electricity_consumption_mwh"], 5.0 * 11630.0, places=1)

    def test_sector_ktoe_split_matches_all_fuels_total(self):
        c = self.out["regions"]["winchester"]["consumption"][str(YEAR)]
        self.assertAlmostEqual(c["sector_ktoe"]["Domestic"], 7.0, places=3)
        self.assertAlmostEqual(c["sector_ktoe"]["Transport"], 8.0, places=3)
        self.assertAlmostEqual(c["sector_ktoe"]["Industrial, Commercial and other"], 6.0, places=3)
        self.assertAlmostEqual(sum(c["sector_ktoe"].values()), c["all_fuels_ktoe"], places=3)

    def test_sector_ktoe_scales_by_retained_fraction_like_fuels_ktoe(self):
        year = str(YEAR)
        expected_domestic = sum(7.0 * lc.MID_HAMPSHIRE_RETAINED_FRACTION[la] for la in lc.MID_HAMPSHIRE_LAS)
        actual_domestic = self.out["regions"]["mid-hampshire"]["consumption"][year]["sector_ktoe"]["Domestic"]
        self.assertAlmostEqual(actual_domestic, expected_domestic, places=3)

    def test_dukes_electricity_mix_parsed_by_year(self):
        mix = self.out["meta"]["dukes_electricity_mix"]
        self.assertAlmostEqual(mix[str(YEAR - 1)]["greenPct"], 19.0, places=3)
        self.assertAlmostEqual(mix[str(YEAR - 1)]["fossilPct"], 81.0, places=3)
        self.assertAlmostEqual(mix[str(YEAR)]["greenPct"], 50.755, places=3)
        self.assertAlmostEqual(mix[str(YEAR)]["fossilPct"], 49.245, places=3)

    def test_dukes_green_and_fossil_pct_sum_to_100(self):
        for entry in self.out["meta"]["dukes_electricity_mix"].values():
            self.assertAlmostEqual(entry["greenPct"] + entry["fossilPct"], 100.0, places=3)

    def test_hampshire_solent_generation_sums_all_las_unweighted(self):
        expected = 150.0 + 65.0 + 15.0 * (len(lc.HAMPSHIRE_SOLENT_LAS) - 2)
        actual = self.out["regions"]["hampshire-solent"]["generation"][str(YEAR)]["total_mwh"]
        self.assertAlmostEqual(actual, expected, places=2)


if __name__ == "__main__":
    unittest.main()
